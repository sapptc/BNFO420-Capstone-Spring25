import pandas as pd
import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Global list to store skipped player summaries (player, reason)
skipped_summary = []

def standardize_position(pos):
    """
    Standardize the position string.
    - All linebacker positions become "LB".
    - All cornerback positions become "CB".
    - Returner positions become "RET".
    - Offensive line positions become "OL".
    - Defensive line positions become "DL".
    - Safety positions become "S".
    - Fullback becomes "FB"
    - Wide receiver becomes "WR"
    - Otherwise, return the original uppercased position.
    """
    pos = str(pos).strip().upper()
    lb_set = {"LB", "OLB", "ILB", "MLB", "WLB", "WILL", "SLB", "SAM", "LILB", "LLB", "ROLB", "LOLB", "RLB", "MILB", "RILB"}
    cb_set = {"CB", "NC", "NCB", "DC", "DCB", "DB", "RCB", "LCB"}
    ret_set = {"RET", "KR", "PR"}
    ol_set = {"T", "OT", "OG", "G", "C", "LG", "RG", "LT", "RT", "LS"}
    dl_set = {"DE", "DT", "NT", "LDT", "RDT", "LDE", "RDE"}
    s_set = {"FS", "SS"}
    fb_set = {"FB"}
    te_set = {"TE"}
    k_set = {"K"}
    p_set = {"P"}
    wr_set = {"WR"}
    rb_set = {"RB"}
    qb_set = {"QB"}
    
    if pos in lb_set:
        return "LB"
    elif pos in cb_set:
        return "CB"
    elif pos in ret_set:
        return "RET"
    elif pos in ol_set:
        return "OL"
    elif pos in dl_set:
        return "DL"
    elif pos in s_set:
        return "S"
    elif pos in fb_set:
        return "FB"    
    elif pos in te_set:
        return "TE"
    elif pos in k_set:
        return "K"
    elif pos in p_set:
        return "P"
    elif pos in wr_set:
        return "WR"
    elif pos in rb_set:
        return "RB"
    elif pos in qb_set:
        return "QB"
    else:
        return pos

def process_file(file_path):
    """
    Process a single .xls file by:
      - Extracting the player's name and reading the file.
      - Converting seasons to numeric and combining duplicate season rows.
      - Determining the player's standardized position.
      - Selecting only the stat columns needed for that position.
      - Writing a CSV output where each row represents a season (kept in year order) without comparing differences.
      - Updating an aggregate Excel file with the season-by-season stats.
    """
    global skipped_summary

    try:
        if not file_path.lower().endswith(".xls"):
            msg = "Wrong format error."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((os.path.basename(file_path), msg))
            return
        if not os.path.exists(file_path):
            msg = "File not found."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((os.path.basename(file_path), msg))
            return

        # Extract player's name from the file name (replace underscores with spaces)
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        player_name = base_filename.replace("_", " ").strip()
        print(f"\nProcessing player: {player_name}")

        # Read the .xls file with header on row 2.
        df = pd.read_excel(file_path, engine='xlrd', header=1)

        # Convert 'Season' column to numeric; drop rows that cannot be converted
        df['Season'] = pd.to_numeric(df['Season'], errors='coerce')
        df = df.dropna(subset=['Season'])
        df['Season'] = df['Season'].astype(int)

        # Verify required columns exist.
        for col in ['Season', 'G', 'Pos']:
            if col not in df.columns:
                msg = f"Required column '{col}' not found."
                print(f"Skipping {file_path}: {msg}")
                skipped_summary.append((player_name, msg))
                return

        # In this modified version, we use all available season data.
        relevant_df = df.copy()

        # Combine duplicate rows for the same season:
        def agg_func(x):
            if x.dtype.kind in 'biufc':
                return x.mean()
            else:
                return x.iloc[0]
        relevant_df = relevant_df.groupby('Season', as_index=False).agg(agg_func)
        relevant_df = relevant_df.sort_values(by='Season')

        # Standardize positions in the relevant data.
        standardized_positions = relevant_df['Pos'].apply(standardize_position)
        unique_positions = standardized_positions.unique()

        # Define the mapping of positions to their required stat columns.
        pos_columns = {
            "QB": ["Yds", "Cmp%", "Int", "TD", "1D"],
            "WR": ["Y/R", "Catch%", "Y/Tgt", "Succ%"],
            "FB": ["Y/R", "Catch%", "Y/Tgt", "Succ%"],
            "RB": ["Y/A", "Y/R", "Succ%", "1D"],
            "TE": ["Y/R", "Catch%", "Y/Tgt", "Succ%"],
            "OL": ["Comb", "Solo", "Ast"],
            "DL": ["PD", "Comb", "Solo", "Ast"],
            "LB": ["PD", "Comb", "Solo", "Ast"],
            "CB": ["PD", "Comb", "Solo", "Ast"],
            "S": ["PD", "Comb", "Solo", "Ast"],
            "K": ["FG%", "Lng"],
            "P": ["Y/P", "Lng"]
        }

        # If more than one standardized position is found, check if the stat sets are identical.
        if len(unique_positions) > 1:
            stat_sets = []
            for pos in unique_positions:
                stats = set([stat for stat in pos_columns.get(pos, []) if stat in df.columns])
                stat_sets.append(stats)
            if all(s == stat_sets[0] for s in stat_sets):
                # Use the position from the latest season record.
                latest_record = relevant_df.iloc[-1]
                pos_from_latest = latest_record['Pos']
                final_pos = standardize_position(pos_from_latest)
                print(f"Multiple positions with identical stat sets found. Categorizing player as position from latest season: {final_pos}")
            else:
                msg = "Position changes with differing stat sets, need further clarification."
                print(f"Skipping {file_path}: {msg}")
                skipped_summary.append((player_name, msg))
                return
        else:
            final_pos = unique_positions[0]
            print(f"Standardized position: {final_pos}")

        # Check if the expected stat columns for the player's position exist.
        stat_cols = [col for col in pos_columns.get(final_pos, []) if col in df.columns]
        if not stat_cols:
            msg = f"Expected stat columns for position {final_pos} not found in file."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((player_name, msg))
            return
        else:
            print(f"Found stat columns for {final_pos}: {stat_cols}")

        # Create result DataFrame containing each season's stats for the required columns.
        # The output will list each season (sorted by year) without any group comparisons.
        result_df = relevant_df[['Season'] + stat_cols].copy()
        result_df.insert(0, "Player", player_name)
        result_df = result_df.sort_values(by='Season')

        # Determine output folder: subfolder in the file's directory named after final_pos.
        file_dir = os.path.dirname(file_path)
        pos_folder = os.path.join(file_dir, final_pos)
        if not os.path.exists(pos_folder):
            os.makedirs(pos_folder)

        # Write individual CSV file.
        output_filename = os.path.join(pos_folder, base_filename + ".csv")
        result_df.to_csv(output_filename, index=False, float_format="%.2f")
        print(f"Individual CSV created: {output_filename}")

        # Aggregate functionality: update aggregate Excel file with per-season data.
        aggregate_filename = os.path.join(pos_folder, f"{final_pos}_aggregate.xlsx")
        if os.path.exists(aggregate_filename):
            try:
                aggregate_df = pd.read_excel(aggregate_filename)
            except Exception as e:
                print(f"Error reading aggregate file: {e}")
                aggregate_df = pd.DataFrame()
        else:
            aggregate_df = pd.DataFrame()

        # Check for duplicate player name and skip if duplicate exists.
        if not aggregate_df.empty and "Player" in aggregate_df.columns and player_name in aggregate_df["Player"].unique():
            msg = "Duplicate player name in aggregate; skipping aggregate update."
            print(f"Skipping aggregate update for {player_name}: {msg}")
            return

        aggregate_df = pd.concat([aggregate_df, result_df], ignore_index=True)
        aggregate_df.to_excel(aggregate_filename, index=False)
        print(f"Aggregate Excel file updated: {aggregate_filename}")

        # Apply alternating row colors (each row now represents one season).
        try:
            wb = load_workbook(aggregate_filename)
            ws = wb.active
            fill1 = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Pale green
            fill2 = PatternFill(start_color="FFFFBD", end_color="FFFFBD", fill_type="solid")  # Pale yellow
            # Data rows start at row 2 (header in row 1)
            for row in range(2, ws.max_row + 1):
                fill = fill1 if (row % 2 == 0) else fill2
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
            wb.save(aggregate_filename)
            print(f"Formatted aggregate Excel file saved: {aggregate_filename}")
        except Exception as e:
            print(f"Error applying formatting: {e}")

    except Exception as ex:
        err_msg = f"An error occurred processing {file_path}: {ex}"
        print(err_msg)
        skipped_summary.append((player_name if 'player_name' in locals() else os.path.basename(file_path), err_msg))

def main_folder():
    global skipped_summary
    folder_path = input("Enter the full path to the folder containing .xls files: ").strip()
    folder_path = folder_path.strip('\'"')
    if not os.path.isdir(folder_path):
        print("Provided folder path does not exist or is not a directory.")
        sys.exit(1)

    file_list = glob.glob(os.path.join(folder_path, "*.xls"))
    if not file_list:
        print("No .xls files found in the provided folder.")
        sys.exit(0)

    print(f"Found {len(file_list)} .xls file(s) in the folder.")
    for file_path in file_list:
        try:
            process_file(file_path)
        except Exception as ex:
            print(f"An error occurred processing {file_path}: {ex}")
            skipped_summary.append((os.path.basename(file_path), str(ex)))

    # At the end of processing, print a summary of skipped files.
    non_duplicate_skips = [item for item in skipped_summary if "Duplicate player name" not in item[1]]
    if non_duplicate_skips:
        print("\nSummary of files not processed:")
        for player, reason in non_duplicate_skips:
            print(f"  {player}: {reason}")
    else:
        print("\nAll files processed successfully.")

if __name__ == "__main__":
    main_folder()
    print("Processing completed.")
    sys.exit(0)