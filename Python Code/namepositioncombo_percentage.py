import pandas as pd
import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Global list to store skipped player summaries (player, reason)
skipped_summary = []

def standardize_position(pos):
    """
    Standardize the position string.
    - All linebacker positions (LB, OLB, ILB, MLB, WLB, WILL, SLB, SAM, LILB, LLB, ROLB, LOLB, RLB, MILB, RILB) become "LB".
    - All cornerback positions (CB, NC, NCB, DC, DCB, DB, RCB, LCB) become "CB".
    - Returner positions (RET, KR, PR) become "RET".
    - Offensive line positions (T, OT, OG, G, C, LG, RG, LT, RT, TE, LS) become "OL".
    - Defensive line positions (DE, DT, NT, LDT, RDT, LDE, RDE) become "DL".
    - Safety positions (FS, SS) become "S".
    - Fullback (FB) becomes "FB"
    - Wide receiver (WR) becomes "WR"
    - Otherwise, return the original uppercased position.
    """
    pos = str(pos).strip().upper()
    lb_set = {"LB", "OLB", "ILB", "MLB", "WLB", "WILL", "SLB", "SAM", "LILB", "LLB", "ROLB", "LOLB", "RLB", "MILB", "RILB"}
    cb_set = {"CB", "NC", "NCB", "DC", "DCB", "DB", "RCB", "LCB"}
    ret_set = {"RET", "KR", "PR"}
    ol_set = {"T", "OT", "OG", "G", "C", "LG", "RG", "LT", "RT", "TE", "LS"}
    dl_set = {"DE", "DT", "NT", "LDT", "RDT", "LDE", "RDE"}
    s_set = {"FS", "SS"}
    fb_set = {"FB"}
    wr_set = {"WR"}
    rb_set = {"RB"}
    qb_set = {"QB"}
    
    if pos in lb_set:
        return "LB"
    elif pos in cb_set:
        return "CB"
    elif pos in ret_set:
        return "RET"
    elif pos in ol_set:
        return "OL"
    elif pos in dl_set:
        return "DL"
    elif pos in s_set:
        return "S"
    elif pos in fb_set:
        return "FB"    
    elif pos in wr_set:
        return "WR"
    elif pos in rb_set:
        return "RB"
    elif pos in qb_set:
        return "QB"
    else:
        return pos

def process_file(file_path):
    """
    Process a single .xls file:
      - Extracts the player's name from the file name.
      - Reads the file and performs validations and standardization.
      - If multiple positions are found, checks if the available stat sets are identical.
         If so, uses the position from the 2024 record.
      - Combines duplicate season rows into one row per year.
      - For group1 (originally 2019, 2020, 2021), if 2019 is missing, substitutes it with the most recent year < 2019.
      - Computes selective stat averages and writes an individual CSV file.
      - Updates the aggregate Excel file for the player's standardized position.
    """
    global skipped_summary

    try:
        if not file_path.lower().endswith(".xls"):
            msg = "Wrong format error."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((os.path.basename(file_path), msg))
            return
        if not os.path.exists(file_path):
            msg = "File not found."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((os.path.basename(file_path), msg))
            return

        # Extract player's name from the file name (replace underscores with spaces)
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        player_name = base_filename.replace("_", " ").strip()
        print(f"\nProcessing player: {player_name}")

        # Read the .xls file with header on row 2.
        df = pd.read_excel(file_path, engine='xlrd', header=1)

        # Convert 'Season' column to numeric; drop rows that cannot be converted
        df['Season'] = pd.to_numeric(df['Season'], errors='coerce')
        df = df.dropna(subset=['Season'])
        df['Season'] = df['Season'].astype(int)

        # Verify required columns exist.
        for col in ['Season', 'G', 'Pos']:
            if col not in df.columns:
                msg = f"Required column '{col}' not found."
                print(f"Skipping {file_path}: {msg}")
                skipped_summary.append((player_name, msg))
                return

        # Define the base required seasons (for group2 we always want 2022-2024)
        base_required = [2022, 2023, 2024]
        # For group1 we expect 2019,2020,2021 but may substitute 2019 if missing.
        group1_base = [2019, 2020, 2021]

        # Filter data for seasons that are either in group1_base or base_required,
        # plus also allow any season less than 2019 as potential substitutes.
        all_years = df['Season'].unique()
        valid_years = set(group1_base + base_required).union({year for year in all_years if year < 2019})
        relevant_df = df[df['Season'].isin(valid_years)]
        if relevant_df.empty:
            msg = "Not enough Seasonal Data."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((player_name, msg))
            return

        # Combine duplicate rows for the same season:
        def agg_func(x):
            if x.dtype.kind in 'biufc':
                return x.mean()
            else:
                return x.iloc[0]
        relevant_df = relevant_df.groupby('Season', as_index=False).agg(agg_func)

        # Determine available seasons after combination.
        available_years = sorted(relevant_df['Season'].unique())
        
        # Check that group2 data exists.
        if not all(year in available_years for year in base_required):
            msg = "Not enough group2 Seasonal Data (2022-2024 missing)."
            print(f"Skipping {file_path}: {msg}")
            skipped_summary.append((player_name, msg))
            return
        
        # For group1, check if 2019 is present.
        if 2019 in available_years:
            group1_years = [2019, 2020, 2021]
        else:
            # Try to find a substitute: the maximum available year less than 2019.
            substitutes = [year for year in available_years if year < 2019]
            if substitutes and all(year in available_years for year in [2020, 2021]):
                sub_year = max(substitutes)
                group1_years = [sub_year, 2020, 2021]
                print(f"Substituting season 2019 with {sub_year} for group1.")
            else:
                msg = "Not enough group1 Seasonal Data (and no suitable substitute for 2019 available)."
                print(f"Skipping {file_path}: {msg}")
                skipped_summary.append((player_name, msg))
                return

        # For each season in group1 and group2, ensure at least one record has G > 6.
        for year in group1_years + base_required:
            season_data = relevant_df[relevant_df['Season'] == year]
            if season_data.empty or season_data['G'].max() < 6:
                msg = f"Not enough games played in season {year}."
                print(f"Skipping {file_path}: {msg}")
                skipped_summary.append((player_name, msg))
                return

        # Define the mapping of positions to the columns (stats) to average.
        pos_columns = {
            "QB": ["Yds", "Cmp%", "Int", "TD", "1D"],
            "WR": ["Y/R", "Catch%", "Y/Tgt", "Succ%"],
            "FB": ["Y/R", "Catch%", "Y/Tgt", "Succ%"],
            "RB": ["Y/A", "Y/R", "Succ%", "1D"],
            "TE": ["Y/R", "Catch%", "Y/Tgt", "Succ%"],
            "OL": ["Comb", "Solo", "Ast"],
            "DL": ["PD", "Comb", "Solo", "Ast"],
            "LB": ["PD", "Comb", "Solo", "Ast"],
            "CB": ["PD", "Comb", "Solo", "Ast"],
            "S": ["PD", "Comb", "Solo", "Ast"],
            "K": ["FG%", "Lng"],
            "P": ["Y/P", "Lng"]
        }

        # Standardize positions in the relevant data.
        standardized_positions = relevant_df['Pos'].apply(standardize_position)
        unique_positions = standardized_positions.unique()

        # If more than one standardized position is found, check stat sets.
        if len(unique_positions) > 1:
            stat_sets = []
            for pos in unique_positions:
                stats = set([stat for stat in pos_columns.get(pos, []) if stat in df.columns])
                stat_sets.append(stats)
            if all(s == stat_sets[0] for s in stat_sets):
                # Use the position from the 2024 record.
                df_2024 = relevant_df[relevant_df['Season'] == 2024]
                if not df_2024.empty:
                    pos_2024 = df_2024['Pos'].iloc[0]
                    final_pos = standardize_position(pos_2024)
                    print(f"Multiple positions with identical stat sets found. Categorizing player as position from 2024: {final_pos}")
                else:
                    msg = "No record for 2024 found to decide position, need further clarification."
                    print(f"Skipping {file_path}: {msg}")
                    skipped_summary.append((player_name, msg))
                    return
            else:
                msg = "Position changes, need further clarification."
                print(f"Skipping {file_path}: {msg}")
                skipped_summary.append((player_name, msg))
                return
        else:
            final_pos = unique_positions[0]
            print(f"Standardized position: {final_pos}")

        # Now we have group1_years (e.g., either [2019,2020,2021] or [substitute,2020,2021])
        # and group2 is always [2022, 2023, 2024].
        group2_years = [2022, 2023, 2024]

        group1 = relevant_df[relevant_df['Season'].isin(group1_years)]
        group2 = relevant_df[relevant_df['Season'].isin(group2_years)]
        avg_group1 = group1[[col for col in pos_columns[final_pos] if col in df.columns]].mean()
        avg_group2 = group2[[col for col in pos_columns[final_pos] if col in df.columns]].mean()
        
        # Compute percentage difference row: ((Group2 - Group1) / Group1) * 100
        diff = ((avg_group2 - avg_group1) / avg_group1) * 100

        # Create result DataFrame with 3 rows.
        result_df = pd.DataFrame([avg_group1, avg_group2, diff],
                                 index=["Group1", "Group2", "Difference"])
        result_df = result_df.reset_index().rename(columns={"index": "Season Group"})
        result_df.insert(0, "Player", player_name)

        # Determine output folder: subfolder in the file's directory named after final_pos.
        file_dir = os.path.dirname(file_path)
        pos_folder = os.path.join(file_dir, final_pos)
        if not os.path.exists(pos_folder):
            os.makedirs(pos_folder)

        # Write individual CSV file with file name as PlayerName_Position.csv (remove spaces in player name)
        output_filename = os.path.join(pos_folder, player_name.replace(" ", "") + "_" + final_pos + ".csv")
        result_df.to_csv(output_filename, index=False, float_format="%.2f")
        print(f"Individual CSV created: {output_filename}")

        # Aggregate functionality:
        aggregate_filename = os.path.join(pos_folder, f"{final_pos}_aggregate.xlsx")
        if os.path.exists(aggregate_filename):
            try:
                aggregate_df = pd.read_excel(aggregate_filename)
            except Exception as e:
                print(f"Error reading aggregate file: {e}")
                aggregate_df = pd.DataFrame()
        else:
            aggregate_df = pd.DataFrame()

        # Check for duplicate player name and skip if duplicate exists (do not log in summary)
        if not aggregate_df.empty and "Player" in aggregate_df.columns and player_name in aggregate_df["Player"].unique():
            msg = "Duplicate player name in aggregate; skipping aggregate update."
            print(f"Skipping aggregate update for {player_name}: {msg}")
            return

        aggregate_df = pd.concat([aggregate_df, result_df], ignore_index=True)
        aggregate_df.to_excel(aggregate_filename, index=False)
        print(f"Aggregate Excel file updated: {aggregate_filename}")

        # Apply alternating row colors by entry (each player's entry = 3 rows).
        try:
            wb = load_workbook(aggregate_filename)
            ws = wb.active
            fill1 = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Pale green
            fill2 = PatternFill(start_color="FFFFBD", end_color="FFFFBD", fill_type="solid")  # Pale yellow
            # Data rows start at row 2 (header in row 1)
            for row in range(2, ws.max_row + 1):
                block_index = (row - 2) // 3
                fill = fill1 if (block_index % 2 == 0) else fill2
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
            wb.save(aggregate_filename)
            print(f"Formatted aggregate Excel file saved: {aggregate_filename}")
        except Exception as e:
            print(f"Error applying formatting: {e}")

    except Exception as ex:
        err_msg = f"An error occurred processing {file_path}: {ex}"
        print(err_msg)
        skipped_summary.append((player_name if 'player_name' in locals() else os.path.basename(file_path), err_msg))

def main_folder():
    global skipped_summary
    folder_path = input("Enter the full path to the folder containing .xls files: ").strip()
    folder_path = folder_path.strip('\'"')
    if not os.path.isdir(folder_path):
        print("Provided folder path does not exist or is not a directory.")
        sys.exit(1)

    file_list = glob.glob(os.path.join(folder_path, "*.xls"))
    if not file_list:
        print("No .xls files found in the provided folder.")
        sys.exit(0)

    print(f"Found {len(file_list)} .xls file(s) in the folder.")
    for file_path in file_list:
        try:
            process_file(file_path)
        except Exception as ex:
            print(f"An error occurred processing {file_path}: {ex}")
            skipped_summary.append((os.path.basename(file_path), str(ex)))

    # At the end of processing, print a summary of skipped players (excluding duplicate errors).
    non_duplicate_skips = [item for item in skipped_summary if "Duplicate player name" not in item[1]]
    if non_duplicate_skips:
        print("\nSummary of files not processed:")
        for player, reason in non_duplicate_skips:
            print(f"  {player}: {reason}")
    else:
        print("\nAll files processed successfully.")

if __name__ == "__main__":
    main_folder()
